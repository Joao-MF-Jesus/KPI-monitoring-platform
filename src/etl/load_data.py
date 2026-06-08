from datetime import date, datetime
from pathlib import Path
import math
import re
import unicodedata

from openpyxl import load_workbook

from src.alerts.rules import check_alerts
from src.database.connection import SessionLocal
from src.database.models import Incident, KPIRecord
from src.kpis.calculate_kpis import calculate_cpa, calculate_cpl, calculate_margem, calculate_roi


def normalize_text(value):
    value = str(value).strip().lower()
    value = unicodedata.normalize("NFKD", value)
    value = value.encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return value.strip()


def to_float(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    value = str(value).strip().replace("R$", "").replace(" ", "")
    if value in ["", "-", "nan", "None"]:
        return 0.0
    if "," in value and "." in value:
        value = value.replace(".", "").replace(",", ".")
    elif "," in value:
        value = value.replace(",", ".")
    try:
        return float(value)
    except ValueError:
        return 0.0


def to_int(value):
    return int(round(to_float(value)))


def find_column_index(headers, options):
    normalized_headers = []
    for index, header in enumerate(headers):
        if header is None:
            continue
        normalized_headers.append((normalize_text(header), index))
    for option in options:
        normalized_option = normalize_text(option)
        for normalized_header, index in normalized_headers:
            if normalized_header == normalized_option:
                return index
    for option in options:
        normalized_option = normalize_text(option)
        for normalized_header, index in normalized_headers:
            if normalized_option in normalized_header:
                return index
    return None


def get_cell(row, index):
    if index is None or index >= len(row):
        return None
    return row[index]


def is_valid_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def extract_records_from_sheet(worksheet):
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration:
        return []
    date_idx = find_column_index(headers, ["Dia"])
    investment_idx = find_column_index(headers, ["Invest. Total", "Investimento"])
    leads_idx = find_column_index(headers, ["Lead - Geral", "Lead"])
    sales_idx = find_column_index(headers, ["Vendas", "Total de Vendas", "Sales"])
    revenue_idx = find_column_index(headers, ["Faturamento Bruto", "Faturamento"])
    profit_idx = find_column_index(headers, ["Lucro Líquido", "Lucro Liquido"])
    if any(index is None for index in [date_idx, investment_idx, leads_idx, sales_idx, revenue_idx, profit_idx]):
        return []
    records = []
    for row in rows:
        row_date = is_valid_date(get_cell(row, date_idx))
        if row_date is None or row_date.year < 2023 or row_date.year > 2026:
            continue
        investimento_ads = to_float(get_cell(row, investment_idx))
        leads = to_int(get_cell(row, leads_idx))
        vendas = to_int(get_cell(row, sales_idx))
        faturamento = to_float(get_cell(row, revenue_idx))
        lucro = to_float(get_cell(row, profit_idx))
        if not any([investimento_ads != 0, leads != 0, vendas != 0, faturamento != 0, lucro != 0]):
            continue
        records.append({
            "data": row_date,
            "source_sheet": worksheet.title,
            "faturamento": faturamento,
            "lucro": lucro,
            "roi": calculate_roi(lucro, investimento_ads),
            "cpa": calculate_cpa(investimento_ads, vendas),
            "cpl": calculate_cpl(investimento_ads, leads),
            "leads": leads,
            "vendas": vendas,
            "margem": calculate_margem(lucro, faturamento),
            "investimento_ads": investimento_ads,
        })
    return records


def extract_excel_records(excel_path):
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Arquivo Excel não encontrado: {excel_path}")
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    all_records = []
    imported_sheets = []
    for worksheet in workbook.worksheets:
        sheet_records = extract_records_from_sheet(worksheet)
        if sheet_records:
            all_records.extend(sheet_records)
            imported_sheets.append((worksheet.title, len(sheet_records)))
    return all_records, imported_sheets


def run_etl(excel_path):
    records, imported_sheets = extract_excel_records(excel_path)
    session = SessionLocal()
    session.query(KPIRecord).delete()
    session.query(Incident).delete()
    session.commit()
    for row in records:
        record = KPIRecord(**row)
        session.add(record)
        session.flush()
        for alert in check_alerts(record):
            session.add(Incident(**alert))
    session.commit()
    session.close()
    print(f"{len(records)} registros importados com sucesso.")
    print("Abas importadas:")
    for sheet_name, total_records in imported_sheets:
        print(f"- {sheet_name}: {total_records} registros")
    return len(records), imported_sheets
